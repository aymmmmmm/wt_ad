#! /usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Thu Mar 14 09:28:48 2019

@author: Minxuan
"""

import pandas as pd
import numpy as np
import pickle
import math
from collections import Counter
import pymongo
import yaml
from decimal import *
import matplotlib.pyplot as plt
import os
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
from openpyxl import Workbook
import openpyxl


wtid= 'turbineName'
time = 'time'
gs = 'generator_rotating_speed'
gp = 'power'
temp_bear_dr = 'f_i_bearing_temp'
temp_bear_ndr = 'r_i_bearing_temp'
temp_cab = 'cabin_temp'
threshold_01 = 0.6
threshold_00 = 0.9
threshold_1 = 1
n_lookback_alarm = 100
n_lookback_warning = 10
n_continuous = 3
border = 0.1


def get_warning_2(res_dr, res_ndr, threshold_01, threshold_00, threshold_1, border):
    """
    param res_dr: 驱动端温度残差序列
    param res_ndr: 非驱动端温度残差序列
    param threshold1: 温度残差预警阈值
    param threshold2: 残差偏移预警阈值
    param border: 判断为异常的比例边界，默认为0.2
    return: 预警值，1表示触发预警
    """
    warning_dr = 0
    warning_ndr = 0
    res_dp = res_dr - res_ndr
    res_dn = res_ndr - res_dr
    index_dr_1 = [i for i, e in enumerate(res_dr) if threshold_01 <= abs(e) < threshold_00]   ###########此处应为绝对值  Jimmy_20211125
    index_ndr_1 = [i for i, e in enumerate(res_ndr) if threshold_01 <= abs(e) < threshold_00]

    index_dr_2 = [i for i, e in enumerate(res_dp) if abs(e) >= threshold_1]
    index_ndr_2 = [i for i, e in enumerate(res_dn) if abs(e) >= threshold_1]

    count_index_dr = [x for x in index_dr_1 if x in index_dr_2]
    count_index_ndr = [x for x in index_ndr_1 if x in index_ndr_2]

    if len(count_index_dr) / len(res_dr) > border or np.sum(abs(res_dr) >= threshold_00) / len(res_dr) > border:  ###【（threshold_01 <= res_dr < threshold_00）且 （res_dp >= threshold_1）】或 【res_dr >= threshold_00】
        warning_dr = 1
    # print(np.sum(abs(res_ndr) >= threshold_00) / len(res_ndr))
    if len(count_index_ndr) / len(res_ndr) > border or np.sum(abs(res_ndr) >= threshold_00) / len(res_ndr) > border:
        # print('ndr_1111')
        warning_ndr = 1
    return warning_dr, warning_ndr

def get_alarm_level(alarm_history, alarm, n_lookback_alarm):
    """
    get_alarm的子函数，根据历史报警积累确定相应的报警等级
    param alarm_series: 历史的报警序列
    param alarm: 当前时刻的报警状态
    return: 报警等级
    """
    alarm_history_tmp = alarm_history[-n_lookback_alarm:]
    # print (alarm_history)
    # distance = min(100 * len([x for x in alarm_history_tmp if x != 0]) / n_lookback_alarm + 40, 100)
    # distance_h = 100 - distance
    # if alarm == 1:  #如果alarm=1才去算距离，如果alarm=0 ，距离就是100， 这个逻辑不太好
    #     if 40 < distance_h <= 60:
    #         alarm_level = 1
    #     elif 20 < distance_h <= 40:
    #         alarm_level = 2
    #     elif distance_h <= 20:
    #         alarm_level = 3

    distance = min(100 * len([x for x in alarm_history_tmp if x != 0]) / n_lookback_alarm , 100)
    distance_h = 100 - distance
    if alarm == 1:  #如果alarm=1才去算距离，如果alarm=0 ，距离就是100， 这个逻辑不太好
        # alarm_level = 0
        if 80 < distance_h <= 100:
            alarm_level = 1
        elif 60 < distance_h <= 80:
            alarm_level = 2
        elif distance_h <= 60:
            alarm_level = 3
    else:
        distance_h = 100
        alarm_level = 0
    return float(distance_h), int(alarm_level)

def get_alarm(warning_dr_history, warning_ndr_history, alarm_dr_history, alarm_ndr_history, n_lookback_alarm, n_lookback_warning, n_continuous):
    """
    根据规则转化为警报，连续n个点或给定时间窗内超过半数预警则产生警报
    param warning_series: 历史的预警序列
    param n_lookback: 每个点的预警考虑前多少个点
    param n_continuous: 连续n个点连续预警
    return： 是否报警
    """
    alarm_dr = 0
    alarm_ndr = 0
    if all(a == 1 for a in warning_dr_history[-n_continuous:]) or find_majority(warning_dr_history[-n_lookback_warning:]) == 1:
        alarm_dr = 1
    if all(a == 1 for a in warning_ndr_history[-n_continuous:]) or find_majority(warning_ndr_history[-n_lookback_warning:]) == 1:
        alarm_ndr = 1
    final_level_dr = get_alarm_level(alarm_dr_history, alarm_dr, n_lookback_alarm)
    final_level_ndr = get_alarm_level(alarm_ndr_history, alarm_ndr, n_lookback_alarm)
    return final_level_dr, final_level_ndr

def data_preprocess(data):
    data = data.sort_values(by = time)
    data['speed'] = data[gs] * math.pi * 2 / 60
    data['espeed'] = (data[gs] * 2 * math.pi / 60) ** 1.6
    data['power'] = abs(data[gp]) * 1000
    data['y_dr'] = data[temp_bear_dr] - 0.5 *(data[temp_cab] + data[temp_cab].shift(1))
    data['y_ndr'] = data[temp_bear_ndr] - 0.5 *(data[temp_cab] + data[temp_cab].shift(1))
    data['temp_cab_mean'] = 0.5 *(data[temp_cab] + data[temp_cab].shift(1))
    data['X1_dr'] = data[temp_bear_dr].shift(1) - 0.5 *(data[temp_cab] + data[temp_cab].shift(1))
    data['X1_ndr'] = data[temp_bear_ndr].shift(1) - 0.5 *(data[temp_cab] + data[temp_cab].shift(1))
    data['X2'] = 0.5 * (data['speed'] + data['speed'].shift(1))
    data['X3'] = 0.5 * (data['espeed'] + data['espeed'].shift(1))
    data['X4'] = 0.5 * (data['power'] + data['power'].shift(1))
    data = data.dropna()
    data.reset_index(inplace=True)
    return data

def find_majority(warning):
    """
    get_warning的子函数
    return: 多数类元素，数量多的那一类
    """
    vote_count = Counter(warning)
    top_two = vote_count.most_common(2)
    if len(top_two) > 1 and top_two[0][1] == top_two[1][1]: # tie
        return 1
    return top_two[0][0]

def import_data_check(data, columns):
    # None
    status_code ='000'
    if data is None: # 判断数据中某列是否全部为空值
        status_code = '301'
        # raise Exception("Input Data_Raw is None")
    # pd.DataFrame()
    elif data.shape[0] == 0:
        # raise Exception('Input Data_Raw is Empty Data_Raw Frame')
        status_code = '300'
        return status_code
    # NAN
    elif data.isnull().all().any():
        # raise Exception('At Least One Column of Input Data_Raw is NAN')
        # for col in columns:
        #     if len(data[data[col].notna()]) == 0:
        #         raise Exception(col + '  is NAN')
        status_code = '300'
        return status_code
    # lack of variable
    if not set(columns).issubset(set(data.columns)):
        # raise Exception('Some Variables are Missing')
        status_code = '300'
        return status_code
    # data type
    # if (data[columns[1:]].dtypes != 'float').any():
    #     # raise Exception('the Dtype of Input Data_Raw is not Float')
    #     status_code = '200'
    #     return status_code
    # duplicate values
    for i in columns[1:]:
        if sum(data[i].duplicated()) == len(data):
            # raise Exception('Input Data_Raw Contains Too Many Duplicates')
            status_code = '101'
    # negative value
    # if any(data[rs] < 0):
    #     # raise Exception('Rotating Speed Contains Negative Values')
    #     status_code = '102'
    # shortage of data
    if (len(data) < 143) & (len(data) >144*0.5):
        # raise Exception('The Volume of Input Data_Raw is Too Low')
        status_code = '101'

    if (len(data) < 144*0.5) & (len(data) >0):
        # raise Exception('The Volume of Input Data_Raw is Too Low')
        status_code = '401'


    return status_code

def generator_temp_main(data,
                        # result_history,
                        model_dr, model_ndr,warning_dr_history,warning_ndr_history,alarm_dr_history,alarm_ndr_history):
    global wtid
    global time
    global gs
    global gp
    global temp_bear_dr
    global temp_bear_ndr
    global temp_cab
    global threshold_01
    global threshold_00
    global threshold_1
    global n_lookback_alarm
    global n_lookback_warning
    global n_continuous
    global model_dr_path
    global model_ndr_path
    global buffer_config_fp

    columns = [time, gs, gp, temp_bear_dr, temp_bear_ndr, temp_cab, wtid]

    data = data[columns]

    # warning_dr_history = []
    # warning_ndr_history = []
    # alarm_dr_history = []
    # alarm_ndr_history = []
    #
    # print (len(warning_dr_history))
    #
    # if len(result_history) < n_lookback_alarm:
    #     for i in range( n_lookback_alarm - len(result_history) ):
    #         warning_dr_history.append(0)
    #         warning_ndr_history.append(0)
    #         alarm_dr_history.append(0)
    #         alarm_ndr_history.append(0)
    #
    # for i in range(len(result_history)):
    #
    #     if 'analysis_data' in result_history.columns:
    #
    #         if (type(result_history.loc[i, 'analysis_data']) == dict ):
    #
    #             warning_dr_history.append(result_history.loc[i, 'analysis_data']['warning_dr'])
    #             warning_ndr_history.append(result_history.loc[i, 'analysis_data']['warning_dr'])
    #         else :
    #             warning_dr_history.append( float('NaN'))
    #             warning_ndr_history.append( float('NaN'))
    #     else:
    #         warning_dr_history.append(float('NaN'))
    #         warning_ndr_history.append(float('NaN'))
    #
    #     if 'gbt1_alarm' in result_history.columns:
    #         alarm_dr_history.append(result_history.loc[i, 'gbt1_alarm'])
    #     else:
    #         alarm_dr_history.append(float('NaN'))
    #
    #     if 'gbt2_alarm' in result_history.columns:
    #         alarm_ndr_history.append(result_history.loc[i, 'gbt2_alarm'])
    #     else:
    #         alarm_ndr_history.append(float('NaN'))


    status_code = import_data_check(data, columns )
    if status_code == '000' or status_code == '101' :
        data_new = data_preprocess(data)

        # data_new.to_csv('temp.csv')

        dr_hat = model_dr.predict(data_new.loc[:,['X1_dr','X2','X3','X4']])
        ndr_hat = model_ndr.predict(data_new.loc[:,['X1_ndr','X2','X3','X4']])
        res_dr = data_new['y_dr'].values.reshape(-1, 1) - dr_hat
        res_ndr = data_new['y_ndr'].values.reshape(-1, 1) - ndr_hat
        res_d = res_dr - res_ndr

        # plt.plot(data_new['y_ndr'].values.reshape(-1, 1))
        # plt.plot(ndr_hat)
        # plt.show()
        #
        # print('np.mean(res_ndr)', np.mean(res_ndr))

        warning = get_warning_2(res_dr, res_ndr, threshold_01, threshold_00, threshold_1, border)
        warning_dr_history.append(warning[0])
        warning_ndr_history.append(warning[1])

        data_new['res_dr'] = res_dr
        data_new['res_ndr']=res_ndr

        #################for test
        # if warning[0] ==1 or warning[1]==1 :
        #
        #
        #     plt.subplot(2, 2, 1)
        #     data_new['dr_hat']=model_dr.predict(data_new.loc[:,['X1_dr','X2','X3','X4']])
        #     data_new['y_dr'].plot(label='(前轴承-机舱)温度_实际')
        #     data_new['dr_hat'].plot(label='(前轴承-机舱)温度_预测')
        #
        #     data_new['res_dr'].plot(label='res')
        #     plt.title(str(data['time'][0])[:-6])
        #
        #     plt.legend()
        #
        #     plt.subplot(2, 2, 3)
        #     data_new['ndr_hat']=model_ndr.predict(data_new.loc[:,['X1_ndr','X2','X3','X4']])
        #     data_new['y_ndr'].plot(label='(后轴承-机舱)温度_实际')
        #     data_new['ndr_hat'].plot(label='(后轴承-机舱)温度_预测')
        #     data_new['res_ndr'].plot(label='res')
        #     plt.title(str(data['time'][0])[:-6])
        #     plt.legend()
        #
        #     plt.subplot(2, 2, 2)
        #     plt.hist(res_dr, bins=100, label='res_前轴承')
        #     plt.title('warning_dr'+ str(warning[0]))
        #     plt.legend()
        #
        #
        #     plt.subplot(2, 2, 4)
        #     plt.hist(res_ndr, bins=100, label='res_后轴承')
        #     plt.title('warning_ndr'+ str(warning[1]))
        #     plt.legend()
        #     plt.subplots_adjust(left=0.04, top=0.96, right=0.96, bottom=0.1, wspace=0.3, hspace=0.3)

            # # plt.show()
            # savePath = '../Result/gen_bearing_temp/{}/'.format(data[wtid][0])
            # if not os.path.exists(savePath):
            #     os.makedirs(savePath)
            # plt.savefig(savePath + str(data['time'][0])[:-6] + 'gen_bearing.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
            # plt.clf()


        alarm = get_alarm(warning_dr_history, warning_ndr_history, alarm_dr_history, alarm_ndr_history, n_lookback_alarm, n_lookback_warning, n_continuous)

        alarm_dr, alarm_ndr = alarm[0], alarm[1]
        alarm_dr_history.append(alarm_dr[1])
        alarm_ndr_history.append(alarm_ndr[1])
        alarm_final = 0
        if np.any([alarm_dr[1], alarm_ndr[1]]) != 0:
            alarm_final = max(alarm_dr[1], alarm_ndr[1])  ####换成两个alarm里面较大的那一个， Jimmy -20211125

        # # plt.show()

        if alarm_final >0:
            savePath = '../Result/gen_bearing_temp/'+ '{}/'.format(data[wtid][0])
            if not os.path.exists(savePath):
                os.makedirs(savePath)

            plt.subplot(2, 2, 1)
            data_new['dr_hat']=model_dr.predict(data_new.loc[:,['X1_dr','X2','X3','X4']])
            data_new['y_dr'].plot(label='(前轴承-机舱)温度_实际')
            data_new['dr_hat'].plot(label='(前轴承-机舱)温度_预测')

            data_new['res_dr'].plot(label='res')
            plt.title(str(data['time'][0])[:-6])

            plt.legend()

            plt.subplot(2, 2, 3)
            data_new['ndr_hat']=model_ndr.predict(data_new.loc[:,['X1_ndr','X2','X3','X4']])
            data_new['y_ndr'].plot(label='(后轴承-机舱)温度_实际')
            data_new['ndr_hat'].plot(label='(后轴承-机舱)温度_预测')
            data_new['res_ndr'].plot(label='res')
            plt.title(str(data['time'][0])[:-6])
            plt.legend()

            plt.subplot(2, 2, 2)
            plt.hist(res_dr, bins=100, label='res_前轴承')
            plt.title('warning_dr'+ str(warning[0]))
            plt.legend()


            plt.subplot(2, 2, 4)
            plt.hist(res_ndr, bins=100, label='res_后轴承')
            plt.title('warning_ndr'+ str(warning[1]))
            plt.legend()
            plt.subplots_adjust(left=0.04, top=0.96, right=0.96, bottom=0.1, wspace=0.3, hspace=0.3)
            plt.savefig(savePath + data[wtid][0]+str(data['time'][0])[:-6] + 'gen_bearing_temp.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
            plt.clf()


        # Construct outputs
        raw_data = {}
        raw_data['datetime'] = list(map(str, data_new[time].tolist()))
        raw_data['gbt1'] = data_new[temp_bear_dr].values.tolist()
        raw_data['gbt2'] = data_new[temp_bear_ndr].values.tolist()
        raw_data['rs'] = data_new[gs].values.tolist()
        raw_data['gp'] = data_new[gp].values.tolist()
        raw_data['gbt1_prediction'] = (dr_hat[:,0] + data_new['temp_cab_mean'].values).tolist()
        raw_data['gbt2_prediction'] = (ndr_hat[:,0] + data_new['temp_cab_mean'].values).tolist()

        analysis_data = {}
        analysis_data['datetime'] = data_new[time].tolist()
        analysis_data['gbt1_residual'] = [float(Decimal(x).quantize(Decimal('0.0000'))) for x in res_dr[:,0].tolist()]
        analysis_data['gbt2_residual'] = [float(Decimal(x).quantize(Decimal('0.0000'))) for x in res_ndr[:,0].tolist()]
        analysis_data['gbt1_residual_average'] = np.mean(res_dr)
        analysis_data['gbt2_residual_average'] = np.mean(res_ndr)
        analysis_data['res_d'] = res_d[:,0].tolist()
        analysis_data['warning_dr'] = warning[0] # Append this to warning_dr_history
        analysis_data['warning_ndr'] = warning[1] # Append this to warning_ndr_history
        analysis_data['gbt_residual_threshold'] = [threshold_00, threshold_01]
        analysis_data['res_d_threshold'] = threshold_1

        result = {}
        result['raw_data'] = raw_data
        result['analysis_data'] = analysis_data
        result['status_code'] = status_code
        result['start_time'] = str(data_new[time].iloc[0])
        result['end_time'] = str(data_new[time].iloc[-1])
        result['gbt1_distance'] = alarm_dr[0]
        result['gbt2_distance'] = alarm_ndr[0]
        result['gbt1_alarm'] = alarm_dr[1] # Append this to alarm_dr_history
        result['gbt2_alarm'] = alarm_ndr[1] # Append this to alarm_ndr_history
        result['distance'] = np.min([alarm_dr[0], alarm_ndr[0]])
        result['alarm'] = alarm_final

    elif status_code == '301' or status_code == '302':

        raw_data = {}
        raw_data['datetime'] = None
        raw_data['gbt1'] = None
        raw_data['gbt2'] = None
        raw_data['rs'] = None
        raw_data['gp'] = None
        raw_data['gbt1_prediction'] = None
        raw_data['gbt2_prediction'] = None

        analysis_data = None

        result = {}
        result['raw_data'] = raw_data
        result['analysis_data'] = analysis_data
        result['status_code'] = status_code
        result['start_time'] = None
        result['end_time'] = None
        result['gbt1_distance'] = None
        result['gbt2_distance'] = None
        result['gbt1_alarm'] = None
        result['gbt2_alarm'] = None
        result['distance'] = None
        result['alarm'] = None



    else:
        raw_data = {}
        raw_data['datetime'] = list(map(str, data[time].tolist()))
        raw_data['gbt1'] = data[temp_bear_dr].values.tolist()
        raw_data['gbt2'] = data[temp_bear_ndr].values.tolist()
        raw_data['rs'] = data[rs].values.tolist()
        raw_data['gp'] = data[gp].values.tolist()
        raw_data['gbt1_prediction'] = None
        raw_data['gbt2_prediction'] = None

        analysis_data = None

        result = {}
        result['raw_data'] = raw_data
        result['analysis_data'] = analysis_data
        result['status_code'] = status_code
        result['start_time'] = str(data[time].iloc[0])
        result['end_time'] = str(data[time].iloc[-1])
        result['gbt1_distance'] = None
        result['gbt2_distance'] = None
        result['gbt1_alarm'] = None
        result['gbt2_alarm'] = None
        result['distance'] = None
        result['alarm'] = None

    # new_result = pd.DataFrame({'raw_data':result['raw_data'],
    #                             'analysis_data':result['analysis_data'] ,
    #                             'status_code':result['status_code'] ,
    #                             'start_time':result['start_time'] ,
    #                             'end_time': result['end_time'] ,
    #                             'gbt1_distance':result['gbt1_distance'] ,
    #                             'gbt2_distance': result['gbt2_distance'] ,
    #                             'gbt1_alarm': result['gbt1_alarm'] ,
    #                             'gbt2_alarm':  result['gbt2_alarm'] ,
    #                             'distance':  result['distance'] ,
    #                             'alarm':  result['alarm'] }, index=['0'])
    #
    # # a= pd.DataFrame(result)
    # # b= pd.DataFrame(new_result)
    # result_history = result_history.append(new_result,ignore_index=True)

    return result['raw_data'],result['analysis_data'] ,result['status_code'],result['start_time'],result['end_time'], result['gbt1_distance'],result['gbt2_distance'],result['gbt1_alarm'], result['gbt2_alarm'], result['alarm'], result['distance']
        # ,result_history

# if __name__ == '__main__':
#     import glob
#     path = r'C:\Project\03_AeroImprovement\05_华润风机SCADA数据分析_202111\华润风机数据\数据\Prep_02_generatorTemp\10min\train_8-11月\3.3MW'
#     allFiles = glob.glob(path + "/*.csv")
#
#     data = pd.DataFrame()
#     list_ = []
#
#     for file_ in allFiles:
#         df = pd.read_csv(file_, engine='python')
#         df.set_index(['time'], inplace=True)
#         df.dropna(inplace=True)
#         list_.append(df)
#
#     data_allturbine = pd.concat(list_)
#
#     savePath = r'C:\Project\03_AeroImprovement\05_华润风机SCADA数据分析_202111\result\generatorBearingTemp\generatorTemp.xlsx'
#     wb = openpyxl.load_workbook(savePath)
#
#
#     for wt in data_allturbine[wtid].unique():
#         data = data_allturbine[data_allturbine[wtid] == wt]
#
#         model_dr_path = '../Resource/dr/{}_generatorBearingTemp_model.bin'.format(wt)
#         model_ndr_path = '../Resource/ndr/{}_generatorBearingTemp_model.bin'.format(wt)
#
#         try:
#             with open(model_dr_path, 'rb') as fp:
#                 model_dr = pickle.load(fp)
#             with open(model_ndr_path, 'rb') as fp:
#                 model_ndr = pickle.load(fp)
#         except:
#             raise Exception('No Such File or Directory')
#
#
#         start_time = '2021-8-9 00:00:00'
#
#         result_history = pd.DataFrame()
#
#         data = data.reset_index(drop=False)
#         data[time] = pd.to_datetime(data[time])
#
#
#         warning_dr_history = []
#         warning_ndr_history = []
#         alarm_dr_history = []
#         alarm_ndr_history = []
#
#
#         rowNum = 1
#         ws = wb.create_sheet(wt)
#
#
#
#         for n in range(90):
#            time_range = pd.date_range(start_time, periods=2, freq='24H')
#            st = time_range[0]
#            et = time_range[1]
#
#            sample_data = data[(data[time] >= st) & (data[time] < et)]
#            sample_data = sample_data.dropna()
#            sample_data.reset_index(inplace=True)
#
#            result ={}
#
#            result['raw_data'],result['analysis_data'] ,result['status_code'],result['start_time'],result['end_time'],\
#            result['gbt1_distance'],result['gbt2_distance'],result['gbt1_alarm'], result['gbt2_alarm'], result['alarm'], result['distance'], result_history \
#                = generator_temp_main(sample_data,result_history,model_dr,model_ndr,warning_dr_history,warning_ndr_history,alarm_dr_history,alarm_ndr_history)
#            # print("result = %s " % result)
#
#            if result['status_code'] == '000' or result['status_code'] == '101' or result['status_code'] == '102' or result['status_code'] == '103':
#                print(wt,st,
#                    ' status_code =', result['status_code'],
#                      'warning_dr =', result['analysis_data']['warning_dr'],
#                      'warning_ndr =', result['analysis_data']['warning_ndr'],
#                      'gbt1_distance =', result['gbt1_distance'],
#                      'gbt2_distance =', result['gbt2_distance'],
#                      'gbt1_alarm =', result['gbt1_alarm'],
#                      'gbt2_alarm =', result['gbt2_alarm'],
#                      'distance =', result['distance'],
#                      'alarm =', result['alarm']
#                      )
#
#                ws.cell(row=rowNum, column=1).value = wt
#                ws.cell(row=rowNum, column=2).value = st
#                ws.cell(row=rowNum, column=3).value = result['status_code']
#                ws.cell(row=rowNum, column=4).value = result['analysis_data']['warning_dr']
#                ws.cell(row=rowNum, column=5).value = result['analysis_data']['warning_ndr']
#                ws.cell(row=rowNum, column=6).value = result['gbt1_distance']
#                ws.cell(row=rowNum, column=7).value = result['gbt2_distance']
#                ws.cell(row=rowNum, column=8).value = result['gbt1_alarm']
#                ws.cell(row=rowNum, column=9).value = result['gbt2_alarm']
#                ws.cell(row=rowNum, column=10).value = result['distance']
#                ws.cell(row=rowNum, column=11).value = result['alarm']
#
#            else:
#                print (wt, st,
#                    ' status_code =', result['status_code'],
#                      'gbt1_distance =', result['gbt1_distance'],
#                      'gbt2_distance =', result['gbt2_distance'],
#                      'gbt1_alarm =', result['gbt1_alarm'],
#                      'gbt2_alarm =', result['gbt2_alarm'],
#                      'distance =', result['distance'],
#                      'alarm =', result['alarm']
#                      )
#
#
#                ws.cell(row=rowNum, column=1).value = wt
#                ws.cell(row=rowNum, column=2).value = st
#                ws.cell(row=rowNum, column=3).value = result['status_code']
#
#                ws.cell(row=rowNum, column=6).value = result['gbt1_distance']
#                ws.cell(row=rowNum, column=7).value = result['gbt2_distance']
#                ws.cell(row=rowNum, column=8).value = result['gbt1_alarm']
#                ws.cell(row=rowNum, column=9).value = result['gbt2_alarm']
#                ws.cell(row=rowNum, column=10).value = result['distance']
#                ws.cell(row=rowNum, column=11).value = result['alarm']
#
#
#
#            rowNum +=1
#
#            start_time = et
#
#         wb.save(savePath)


if __name__ == '__main__':
    import glob
    path = r'C:\Project\03_AeroImprovement\05_华润风机SCADA数据分析_202111\华润风机数据\数据\Prep_02_generatorTemp\10min\origin_10min'
    allFiles = glob.glob(path + "/*.csv")

    data = pd.DataFrame()
    list_ = []

    for file_ in allFiles:
        df = pd.read_csv(file_, engine='python')
        df.set_index(['time'], inplace=True)
        df.dropna(inplace=True)
        list_.append(df)

    data = pd.concat(list_)
    i =1
    for wt in data[wtid].unique():
        df = data[data[wtid] == wt]

        #
        # model_dr_path = './Resource_old/dr/WT{}_model.bin'.format(wtid)
        # model_ndr_path = './Resource_old/ndr/WT{}_model.bin'.format(wtid)

        model_dr_path = '../Resource/dr/{}_generatorBearingTemp_model.bin'.format(wt)
        model_ndr_path = '../Resource/ndr/{}_generatorBearingTemp_model.bin'.format(wt)

        try:
            with open(model_dr_path, 'rb') as fp:
                model_dr = pickle.load(fp)
            with open(model_ndr_path, 'rb') as fp:
                model_ndr = pickle.load(fp)
        except:
            raise Exception('No Such File or Directory')

        data_new = data_preprocess(df)

        data_new['dr_hat'] = model_dr.predict(data_new.loc[:, ['X1_dr', 'X2', 'X3', 'X4']])
        data_new['ndr_hat'] = model_ndr.predict(data_new.loc[:, ['X1_ndr', 'X2', 'X3', 'X4']])
        data_new['res_dr'] = data_new['y_dr'] - data_new['dr_hat']
        data_new['res_ndr'] = data_new['y_ndr']- data_new['ndr_hat']
        data_new['res_d'] = data_new['res_dr'] - data_new['res_ndr']

        data_new = data_new[(data_new['res_dr']<3 ) & (data_new['res_dr']>-3 )]
        data_new = data_new[(data_new['res_ndr'] < 3) & (data_new['res_ndr'] > -3)]


        plt.subplot(5, 5, i)
        i += 1

        # plt.hist(data_new['res_dr'], bins=100, density=True, label='(前轴承-机舱)温度_残差分布')

        # print (wtid, np.mean(data_new['res_dr']), np.std(data_new['res_dr']), 3*np.std(data_new['res_dr']))
        # print(wt, np.mean(data_new['res_ndr']), np.std(data_new['res_ndr']), 3 * np.std(data_new['res_ndr']))
        # plt.hist(data_new['res_ndr'], bins=100, density=True, label='(后轴承-机舱)温度_残差分布')
        # data_new['res_dr'].plot(label='(前轴承-机舱)温度预测残差')

        # plt.scatter(data_new.index, data_new['res_dr'], label='(前轴承-机舱)温度预测残差', s=0.5)

        # plt.scatter(data_new.index, data_new['res_ndr'], label='(后轴承-机舱)温度预测残差', s=0.5)

        # data_new['y_dr'].plot(label='(前轴承-机舱)温度_实际')
        # data_new['dr_hat'].plot(label='(前轴承-机舱)温度_预测')
        # data_new['y_ndr'].plot(label='(后轴承-机舱)温度_实际')
        # data_new['ndr_hat'].plot(label='(后轴承-机舱)温度_预测')
        # plt.ylim(-10, 40)
        # plt.ylim(-2,2)
        plt.xlim(-2, 2)
        plt.legend()
        plt.grid()
        # plt.ylabel('/℃')
        # plt.xlabel('/rpm')
        # plt.xlabel('/kW')
        # plt.ylabel('Density')
        # plt.xlabel('温度/℃')
        plt.title(wt, fontsize=15)
        plt.subplots_adjust(left=0.04, top=0.96, right=0.96, bottom=0.1, wspace=0.01, hspace=0.3)
        # plt.show()
        # generator_bearing_fsrc_main(df,wtid)
    plt.show()







