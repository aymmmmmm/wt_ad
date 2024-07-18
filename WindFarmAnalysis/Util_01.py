# -*- coding: utf-8 -*-
# @Time    : 2021/7/29 17:46
# @File    : PowerCurveProcess0729.py
# @Software: PyCharm
import pandas as pd
import numpy as np
from scipy.stats import weibull_min
import math
import openpyxl
import os
import os.path
import seaborn
import matplotlib.pyplot as plt
import warnings
# action参数可以设置为ignore，once表示为只显示一次
warnings.filterwarnings(action='ignore')
import scipy
from scipy.stats import weibull_min
from scipy.optimize import curve_fit
from statsmodels.base.model import GenericLikelihoodModel
import datetime
from dateutil.relativedelta import relativedelta
from matplotlib.pyplot import MultipleLocator
from matplotlib.dates import DayLocator
import xlwt
import xlsxwriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from windrose import WindroseAxes
import matplotlib.cm as cm
from windrose import plot_windrose


ye= 'wind_direction'
ws= 'wind_speed'
gp = 'power'
gs= 'rotor_speed'
vib_x= 'nacelle_acc_X'
vib_y= 'nacelle_acc_Y'
yaw= 'yaw_error'
ts= 'time'
bp= 'pitch_1'
gs= 'generator_rotating_speed'

rs_lower = 6
wsin =3
gp_lower = 10
maxgp = 3200
rated_rs = 12
gp_percent_limit =0.95
ba_limit = 1
ba_upper = 90

plot_hight=3



class powerCurve:

    turbineName =''
    list_eachbin_df = [] #每一个bin里所有的数据,dataFrame 的 list
    eachbin_stats_data = pd.DataFrame(columns= ['V_average_processed','P_average_processed', 'V_l', 'V_u', 'V_label','vib_magn_average_processed','lambda_ave','Cp_ave', 'windDirection_ave']) #DataFrame, 行数为bin的数量

    def printValue(self):
        for i in range(len(self.eachbin_stats_data)):
            print (self.eachbin_stats_data.loc[i,'V_average_processed'], self.eachbin_stats_data.loc[i,'P_average_processed'],
                   'V_label' , self.eachbin_stats_data.loc[i,'V_label'])

    def print_to_excel(self, savePath, sheet, columnStart,delta_bin ,V_in, V_out):
        wb = openpyxl.load_workbook(savePath)
        ws = wb[sheet]
        ws.cell(row=2, column=columnStart+2).value = self.turbineName
        ws.cell(row=5+40, column=columnStart + 2).value = self.turbineName

        bin_Num = int((V_out - V_in) / delta_bin)
        windSpeedLables = np.linspace(V_in + delta_bin / 2, V_out - delta_bin / 2, bin_Num)

        for n in range(len(windSpeedLables)):
            ws.cell(row=n + 3, column=2).value = windSpeedLables[n]
            ws.cell(row=6 + 40 + n, column=2).value = windSpeedLables[n]

        for i in range(len(self.list_eachbin_df)):

            for n in range(len(windSpeedLables)):
                if (windSpeedLables[n]==self.eachbin_stats_data.loc[i,'V_label']):
                    ws.cell(row=n + 3, column=columnStart+2).value = self.eachbin_stats_data.loc[i,'P_average_processed']
                    ws.cell(row=6+40+n, column=columnStart + 2).value = self.eachbin_stats_data.loc[i, 'Cp_ave']
        wb.save(savePath)

    def print_to_excel_byMonth(self, savePath, sheet, columnStart, startMonth,delta_bin ,V_in, V_out):
        wb = openpyxl.load_workbook(savePath)
        ws = wb[sheet]
        ws.cell(row=1, column=columnStart+2).value = self.turbineName
        ws.cell(row=2, column=columnStart + 2).value = startMonth
        ws.cell(row=5+39, column=columnStart + 2).value = self.turbineName
        ws.cell(row=5 + 40, column=columnStart + 2).value = startMonth

        bin_Num = int((V_out - V_in) / delta_bin)
        windSpeedLables = np.linspace(V_in + delta_bin / 2, V_out - delta_bin / 2, bin_Num)

        for n in range(len(windSpeedLables)):
            ws.cell(row=n + 3, column=2).value = windSpeedLables[n]
            ws.cell(row=6 + 40 + n, column=2).value = windSpeedLables[n]

        for i in range(len(self.list_eachbin_df)):
            for n in range(len(windSpeedLables)):
                if (windSpeedLables[n] == self.eachbin_stats_data.loc[i, 'V_label']):
                    ws.cell(row=n + 3, column=columnStart + 2).value = self.eachbin_stats_data.loc[i, 'P_average_processed']
                    ws.cell(row=6 + 40 + n, column=columnStart + 2).value = self.eachbin_stats_data.loc[i, 'Cp_ave']
        wb.save(savePath)

class vibCurve:

    turbineName =''
    list_eachbin_df = [] #每一个bin里所有的数据,dataFrame 的 list
    eachbin_stats_data = pd.DataFrame(columns= ['V_average_processed', 'V_l', 'V_u', 'V_label','vib_magn_average_processed']) #DataFrame, 行数为bin的数量

    def printValue(self):
        for i in range(len(self.eachbin_stats_data)):
            print (self.eachbin_stats_data.loc[i,'V_average_processed'], self.eachbin_stats_data.loc[i,'vib_magn_average_processed'],
                   'V_label' , self.eachbin_stats_data.loc[i,'V_label'])


    def print_to_excel(self, savePath, sheet, columnStart,delta_bin ,V_in, V_out):
        wb = openpyxl.load_workbook(savePath)
        ws = wb[sheet]
        ws.cell(row=2, column=columnStart+2).value = self.turbineName
        ws.cell(row=5+40, column=columnStart + 2).value = self.turbineName

        bin_Num = int((V_out - V_in) / delta_bin)
        windSpeedLables = np.linspace(V_in + delta_bin / 2, V_out - delta_bin / 2, bin_Num)

        for n in range(len(windSpeedLables)):
            ws.cell(row=n + 3, column=2).value = windSpeedLables[n]
            ws.cell(row=6 + 40 + n, column=2).value = windSpeedLables[n]

        for i in range(len(self.list_eachbin_df)):

            for n in range(len(windSpeedLables)):
                if (windSpeedLables[n]==self.eachbin_stats_data.loc[i,'V_label']):
                    ws.cell(row=n + 3, column=columnStart+2).value = self.eachbin_stats_data.loc[i,'vib_magn_average_processed']
        wb.save(savePath)

def reLoad_ScadaData (pre_Dir, neg_number):
    fileNames = pd.DataFrame(columns=['filePath','turbine_name'])
    for root, dirs, files in os.walk(pre_Dir, topdown=False):
        for name in files:
            fileNames =fileNames.append ([{'filePath':os.path.join(root, name),'turbine_name':name.split('.')[0]}], ignore_index=True)

    list_df_preProcessed_WTs = []
    for i in range (len(fileNames)-neg_number):
        df_scada_WT_preProcessed_one = pd.read_csv(fileNames.loc[i,'filePath'],encoding='gb2312',parse_dates=['time'],infer_datetime_format=True, index_col=0)
        df_scada_WT_preProcessed_one = df_scada_WT_preProcessed_one.drop_duplicates([ts])

        # print (df_scada_WT_preProcessed_one.columns)
        for colu in df_scada_WT_preProcessed_one.columns:
            print (fileNames.loc[i,'turbine_name'], colu, 'has' , df_scada_WT_preProcessed_one[colu].isnull().sum(), 'Null Values in', len(df_scada_WT_preProcessed_one))
        print('Loaded', fileNames.loc[i, 'turbine_name'], 'pointNumber=', len(df_scada_WT_preProcessed_one))

        df_scada_WT_preProcessed_one['turbineName']=fileNames.loc[i, 'turbine_name']
        # plt.scatter(df_scada_WT_preProcessed_one['wind_speed'], df_scada_WT_preProcessed_one['wind_direction'], s=2)
        # plt.show()
##################临时处理

        if ('pressure' not in df_scada_WT_preProcessed_one.columns):
            df_scada_WT_preProcessed_one['pressure'] = 101325
        if ('temperature' not in df_scada_WT_preProcessed_one.columns):
            df_scada_WT_preProcessed_one['temperature'] = df_scada_WT_preProcessed_one['environment_temp']
        if ('humidity' not in df_scada_WT_preProcessed_one.columns):
            df_scada_WT_preProcessed_one['humidity'] = 100

        list_df_preProcessed_WTs.append(df_scada_WT_preProcessed_one)
    return list_df_preProcessed_WTs

# Index(['time', 'turbineName', 'windSpeed', 'windDirection', 'nacellePosition',
#        'yawError', 'power', 'rotorSpeed', 'generatorSpeed', 'pitch_1',
#        'pitch_2', 'pitch_3', 'nacelleVib_X', 'nacelleVib_Y', 'temperature',
#        'pressure', 'humidity'],

def onlyWorking (list_df_preProcessed_WTs, min_rotorSpeed):
    for i in range(len(list_df_preProcessed_WTs)):
        list_df_preProcessed_WTs[i]= list_df_preProcessed_WTs[i][(list_df_preProcessed_WTs[i][ws]>3) & (list_df_preProcessed_WTs[i][gp] > 5) & (list_df_preProcessed_WTs[i][gs] > min_rotorSpeed)]
    return list_df_preProcessed_WTs


def dropNull (list_df_preProcessed_WTs, dropWindSpeed, dropPower, dropRotorSpeed, dropVib, dropTemp):
    for i in range(len(list_df_preProcessed_WTs)):
        if (dropWindSpeed==1):
            list_df_preProcessed_WTs[i].dropna(subset=[ws],inplace=True)
        if (dropPower==1):
            list_df_preProcessed_WTs[i].dropna(subset=[gp],inplace=True)
        if (dropVib==1):
            list_df_preProcessed_WTs[i].dropna(subset=[vib_x],inplace=True)
            list_df_preProcessed_WTs[i].dropna(subset=[vib_y], inplace=True)
        if (dropRotorSpeed==1):
            list_df_preProcessed_WTs[i].dropna(subset=[gs], inplace=True)
        if (dropTemp==1):
            list_df_preProcessed_WTs[i].dropna(subset=['temperature'],inplace=True)
            # list_df_preProcessed_WTs[i].dropna(subset=['pressure'], inplace=True)

        list_df_preProcessed_WTs[i]=list_df_preProcessed_WTs[i].reset_index(drop=True)
        print(list_df_preProcessed_WTs[i].loc[0]['turbineName'], 'Dropped to ', len(list_df_preProcessed_WTs[i]))
    return list_df_preProcessed_WTs

def origin_statas_windSpeed (list_df_WTs, subNumber_H, subNumber_W, cut_lead, cut_tail, output_Dir, ifShow):
    print ('Wait for plotting wind distribution...')
    savePath = output_Dir
    if not os.path.exists(savePath):
        os.makedirs(savePath)
    f = open(savePath + 'MeanWindSpeed.txt', 'w')
    figureHeight = plot_hight* math.ceil((len(list_df_WTs)-cut_tail- cut_lead)/5)
    plt.figure(figsize=(16, figureHeight))
    for i in range (cut_lead, (len(list_df_WTs)-cut_tail)):
        list_df_WTs[i] = list_df_WTs[i][list_df_WTs[i][ws] < 30.0]  # 删除风速>30的项

        if np.nanmax(list_df_WTs[i][yaw]) > 180:
            list_df_WTs[i].loc[list_df_WTs[i][yaw] >= 180, yaw] = list_df_WTs[i].loc[list_df_WTs[i][yaw] >= 180, yaw] - 360

        list_df_WTs[i]['windSpeed_x']= list_df_WTs[i][ws]*np.cos(list_df_WTs[i][yaw]*np.pi/180.0)
        list_df_WTs[i]['windSpeed_y'] = list_df_WTs[i][ws] * np.sin(list_df_WTs[i][yaw] * np.pi / 180.0)

        print (str(list_df_WTs[i].loc[0]['turbineName']), 'mean' , np.mean(list_df_WTs[i][ws]), 'std', np.std(list_df_WTs[i][ws]))
        # print(str(list_df_WTs[i].loc[0]['turbineName']), 'mean', np.mean(list_df_WTs[i]['windSpeed_x']), 'std', np.std(list_df_WTs[i]['windSpeed_x']))
        # print(str(list_df_WTs[i].loc[0]['turbineName']), 'mean', np.mean(list_df_WTs[i]['windSpeed_y']), 'std', np.std(list_df_WTs[i]['windSpeed_y']))


        plt.subplot(subNumber_H,subNumber_W,i+1) # For 全部显示
        seaborn.distplot(list_df_WTs[i][ws],axlabel='')

        x_major_locator = MultipleLocator(5)
        y_major_locator = MultipleLocator(0.1)
        ax = plt.gca()
        ax.xaxis.set_major_locator(x_major_locator)
        ax.yaxis.set_major_locator(y_major_locator)


        plt.xlim((0, 20))
        plt.grid()
        plt.ylim((0,0.3))
        plt.title(str(list_df_WTs[i].loc[0]['turbineName']) + '_V_ave=' + str(round(np.mean (list_df_WTs[i][ws]),2)) , fontsize='large',y=0.5)
        plt.tick_params(labelsize=10)
        plt.subplots_adjust(left=0.04, top=0.96, right=0.96, bottom=0.1, wspace=0.2, hspace=0.3)
        f.write('V_average:  '+ str(list_df_WTs[i].loc[0]['turbineName']) + '   ' + str(round(np.mean(list_df_WTs[i][ws]), 2))+ ' turbulence =' + str(round(np.std(list_df_WTs[i][ws])/np.mean(list_df_WTs[i][ws]),2))  +'\n')
    # plt.legend(loc="best")
    f.close()
    plt.savefig(savePath+ 'windDistribution.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
    if (ifShow==True):
        plt.show()
    plt.clf()



def origin_statas_wind_rose (list_df_WTs, subNumber_H, subNumber_W, cut_lead, cut_tail, output_Dir, ifShow):
    print('Wait for plotting wind rose map...')
    savePath = output_Dir
    # plt.figure(figsize=(16, 9))
    figureHeight = plot_hight * math.ceil((len(list_df_WTs) - cut_tail - cut_lead) / 5)
    fig = plt.figure(figsize=(16, figureHeight))
    for i in range(cut_lead, (len(list_df_WTs) - cut_tail)):
        list_df_WTs[i] = list_df_WTs[i][list_df_WTs[i][ws] < 30.0]  # 删除风速>30的项

        if np.nanmax(list_df_WTs[i][yaw]) > 180:
            list_df_WTs[i].loc[list_df_WTs[i][yaw] >= 180, yaw] = list_df_WTs[i].loc[list_df_WTs[i][yaw] >= 180, yaw] - 360

        ax = fig.add_subplot(subNumber_H, subNumber_W, i + 1, projection="windrose")
        ax.bar(list_df_WTs[i][ye], list_df_WTs[i][ws], normed=True, opening=0.8, edgecolor='white')
        # ax.set_legend(fontsize=1)
        ax.set_title(str(list_df_WTs[i].loc[0]['turbineName']), fontsize=10,loc='left' )

        plt.subplots_adjust(left=0.04, top=0.96, right=0.96, bottom=0.1, wspace=0.2, hspace=0.3)

    plt.savefig(savePath + 'wind_rose.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
    if (ifShow == True):
        plt.show()
    plt.clf()

def origin_statas_wind_corelation (list_df_WTs, subNumber_H, subNumber_W, cut_lead, cut_tail, output_Dir, ifShow):
    data_turbine = list_df_WTs[0][['time', 'wind_speed']]
    # data_turbine['time'] = data_turbine['time'].map(lambda x: datetime.datetime.strftime(x, '%Y-%m-%d %H:%M:%S'))
    data_turbine.columns=['time', str(list_df_WTs[0].loc[0]['turbineName'])]
    data_turbine = data_turbine.drop_duplicates(['time'])
    for i in range(1, len(list_df_WTs)):
        temp = list_df_WTs[i][['time', 'wind_speed']]
        # temp['time'] =temp['time'].map(lambda x: datetime.datetime.strftime(x, '%Y-%m-%d %H:%M:%S'))
        temp.columns = ['time', str(list_df_WTs[i].loc[0]['turbineName'])]
        temp = temp.drop_duplicates(['time'])   ###需要去重，否则merge 会出现很多重复值

        data_turbine = pd.merge(data_turbine, temp, on=['time'], how= 'inner')
        print (str(list_df_WTs[i].loc[0]['turbineName']), len(temp), len(data_turbine))

    # data_turbine.to_csv(output_Dir+'merge.csv',encoding='gbk' )
    correlation= data_turbine.corr(method='pearson')
    correlation.to_csv(output_Dir + 'wind_correlation.csv', encoding='gbk')

    pass




def origin_statas (list_df_WTs, variable_x, variable_y, variable_y2, power_lowbound, rotorSpeed_lowbound, ifpreprocess, cut_lead,cut_tail, output_Dir, ifShow,subNumber_H, subNumber_W):
    print ('Wait for plotting statas...'+ variable_x +'-'+ variable_y)
    # savePath = output_Dir + str(list_df_WTs[i].loc[0]['turbineName']) + r'\\'
    savePath = output_Dir
    if not os.path.exists(savePath):
        os.makedirs(savePath)
    # start_time = datetime.datetime.strptime('2021-1-1 0:00:00', '%Y-%m-%d %H:%M:%S')
    figureHeight = plot_hight * math.ceil((len(list_df_WTs) - cut_tail - cut_lead) / 5)
    plt.figure(figsize=(16, figureHeight))
    for i in range (cut_lead, (len(list_df_WTs)-cut_tail)):
        # list_df_WTs[i]=list_df_WTs[i][list_df_WTs[i]['time']>start_time]
        # list_df_WTs[i] = list_df_WTs[i].reset_index(drop=True)
        if (ifpreprocess==1):
            list_df_WTs[i] = list_df_WTs[i][list_df_WTs[i][gp] > power_lowbound]  # 删除功率<=power_lowbound的行
            list_df_WTs[i] = list_df_WTs[i][list_df_WTs[i][gs] > rotorSpeed_lowbound]  # 删除转速<=rotorSpeed_lowbound的行
            list_df_WTs[i]=list_df_WTs[i].reset_index(drop=True)
#####################################

        plt.subplot(subNumber_H,subNumber_W,i+1)

        # plt.figure(i)      # For 单独显示
        plt.scatter(list_df_WTs[i][variable_x], list_df_WTs[i][variable_y], s=0.1, color='blue')
        if (len(variable_y2)>0 ):
            plt.scatter(list_df_WTs[i][variable_x], list_df_WTs[i][variable_y2], color='r', s=0.1)

        ####设置x轴刻度
        if(variable_x==ws):
            plt.xlim((0, 20))
            x_major_locator = MultipleLocator(5)
        if (variable_x == 'windSpeed_Standard'):
            plt.xlim((0, 20))
            x_major_locator = MultipleLocator(5)
        if (variable_x == 'windSpeed_yawCorrected'):
            plt.xlim((0, 20))
            x_major_locator = MultipleLocator(5)
        if (variable_x == gs):
            plt.xlim((5, rated_rs+3))
            x_major_locator = MultipleLocator(5)
        if (variable_x == ts):
            x_major_locator = MultipleLocator(100)
        if (variable_x == gp):
            plt.xlim((0, maxgp+300))
            x_major_locator = MultipleLocator(500)
        if (variable_x == ye):
            plt.xlim((-50, 50))
            x_major_locator = MultipleLocator(10)
        if (variable_x == 'lambda'):
            plt.xlim((0, 20))
            x_major_locator = MultipleLocator(5)




####设置y轴刻度
        if (variable_y == gp):
            y_major_locator = MultipleLocator(500)
            plt.ylim((-30, maxgp+300))
        if (variable_y == gs):
            y_major_locator = MultipleLocator(2)
            plt.ylim((0, rated_rs+3))
            y_major_locator = MultipleLocator(2)
        if (variable_y == bp):
            y_major_locator = MultipleLocator(10)
        if (variable_y == 'lambda'):
            y_major_locator = MultipleLocator(5)
        if (variable_y == yaw):
            y_major_locator = MultipleLocator(10)
        if (variable_y == vib_x):
            y_major_locator = MultipleLocator(0.05)
            plt.ylim((-0.02, 0.5))
        if (variable_y == vib_y):
            y_major_locator = MultipleLocator(0.05)
            plt.ylim((-0.02, 0.5))
        if (variable_y == 'nacelleVib_X_new'):
            y_major_locator = MultipleLocator(0.05)
            plt.ylim((-0.02, 0.5))
        if (variable_y == 'nacelleVib_Y_new'):
            y_major_locator = MultipleLocator(0.05)
            plt.ylim((-0.02, 0.5))
        if (variable_y == 'vib_magn_new'):
            y_major_locator = MultipleLocator(0.02)
            plt.ylim((-0.02, 0.5))
        if (variable_y == 'temperature'):
            y_major_locator = MultipleLocator(5)
            plt.ylabel('T ℃')
        if (variable_y == 'pressure'):
            y_major_locator = MultipleLocator(1000)
        if (variable_y == 'humidity'):
            y_major_locator = MultipleLocator(0.2)
        if (variable_y == 'Rho_standard'):
            y_major_locator = MultipleLocator(0.05)
        if (variable_y == 'Ps'):
            y_major_locator = MultipleLocator(1000)
        if (variable_y == 'Cp'):
            y_major_locator = MultipleLocator(0.2)
            plt.ylim((0, 1.4))
        if (variable_y == 'Cp_noRhoCorrect'):
            y_major_locator = MultipleLocator(0.2)
            plt.ylim((0, 1.4))
        if (variable_y == ye):
            y_major_locator = MultipleLocator(10)
            plt.ylim((-30, 30))
        if(variable_y == yaw):
            plt.ylim((-30, 30))
            y_major_locator = MultipleLocator(10)
        if (variable_y == ws):
            plt.ylim((0, 20))
            y_major_locator = MultipleLocator(5)
        if (variable_y == 'torque_measure'):
            plt.ylim((0, 20000))
            y_major_locator = MultipleLocator(3000)
        if (variable_y == bp):
            plt.ylim((0, 25))
            y_major_locator = MultipleLocator(5)

        ax=plt.gca()
        ax.xaxis.set_major_locator(x_major_locator)
        ax.yaxis.set_major_locator(y_major_locator)
        # plt.ylim((list_df_WTs[0][variable_y].min(),list_df_WTs[0][variable_y].max()*1.1))


        # if (variable_x == 'time'):
        #     plt.xticks(pd.date_range('20210101', '20211101', freq='2M'),  rotation= 30)

        plt.xlabel(variable_x)
        plt.ylabel(variable_y)
        plt.grid(linestyle='--',which='major')
        plt.title('Turbine_' + str(list_df_WTs[i].loc[0]['turbineName']), fontsize='large',y=0.1,verticalalignment='top', horizontalalignment='left' )
        plt.tick_params(labelsize=10)
        plt.subplots_adjust(left=0.2, top=0.96, right=0.96, bottom=0.1, wspace=0.3, hspace=0.3)
        plt.margins(0, 0)


    # plt.legend(loc="best")

    plt.savefig(savePath+ variable_x +'  '+ variable_y +'.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
    if (ifShow==True):
        plt.show()
    plt.clf()

def TSR(W,U,R):
    try:
        tsr= W * math.pi / 30 * R / U
    except:
        tsr='NA'
    return tsr

def cal_TSR(list_df_WTs,R):
    for i in range(len(list_df_WTs)):
        list_df_WTs[i]['lambda'] = list_df_WTs[i][gs] * np.pi / 30 * R / list_df_WTs[i][ws]

def virb_magn(x, y):
    a = x * x + y * y
    return (math.pow(a, 0.5))

def vib_shift_X (df_scada_WT_1):
    df_scada_WT=df_scada_WT_1.copy()
    delta_bin = 0.2
    V_in = 1
    V_out = 2.5
    bin_Num = int((V_out - V_in) / delta_bin)
    windSpeedBins = np.linspace(V_in, V_out, bin_Num + 1)
    windSpeedLables = np.linspace(V_in + delta_bin / 2, V_out - delta_bin / 2, bin_Num)
    df_scada_WT['windSpeed_lable'] = pd.cut(df_scada_WT[ws], bins=windSpeedBins, labels=windSpeedLables,   right=False)  # 按风速bin分组
    df_scada_WT = df_scada_WT.dropna()  # 删除无法归到bin里的空值
    df_scada_WT = df_scada_WT.reset_index(drop=True)
    df_scada_WT_grouped = df_scada_WT.groupby(df_scada_WT['windSpeed_lable'])  # 按风速bin分组
    powerCurve01 = powerCurve()
    powerCurve01.turbineName = df_scada_WT.loc[0, 'turbineName']
    list_eachbin_df = []
    for name, group in df_scada_WT_grouped:
        if (len(group) > 0):
            df_temp = df_scada_WT_grouped.get_group(name)  # get_group 才能返回dataFrame, group 是tuple
            list_eachbin_df.append(df_temp)

    nacellVib_X_ave = pd.DataFrame()
    for i_bin in range(len(list_eachbin_df)):  # 对每一个风速bin，进行数据清洗 , power 四分位法
        res_var = np.percentile(list_eachbin_df[i_bin]['nacelleVib_X'], (25, 50, 75), interpolation='midpoint')
        Iqr_var = res_var[2] - res_var[0]
        var_l = res_var[0] - 0.001 * Iqr_var
        var_u = res_var[2] + 0.001 * Iqr_var
        if (len(list_eachbin_df[i_bin]) > 5):
            list_eachbin_df[i_bin] = list_eachbin_df[i_bin][
                (list_eachbin_df[i_bin][vib_x] >= var_l) & (list_eachbin_df[i_bin][vib_x] <= var_u)]
        new = pd.DataFrame({'windSpeed_ave': np.mean(list_eachbin_df[i_bin][ws]), 'X_ave':np.mean(list_eachbin_df[i_bin][vib_x])}, index = ['0'])
        nacellVib_X_ave = nacellVib_X_ave.append(new,ignore_index=True)
    df_scada_WT_1['nacellVib_X_shift']=np.mean(nacellVib_X_ave['X_ave'])


def vib_shift_Y (df_scada_WT_1):

    df_scada_WT=df_scada_WT_1.copy()
    delta_bin = 0.2
    V_in = 1
    V_out = 2.5

    bin_Num = int((V_out - V_in) / delta_bin)
    windSpeedBins = np.linspace(V_in, V_out, bin_Num + 1)
    windSpeedLables = np.linspace(V_in + delta_bin / 2, V_out - delta_bin / 2, bin_Num)


    df_scada_WT['windSpeed_lable'] = pd.cut(df_scada_WT[ws], bins=windSpeedBins, labels=windSpeedLables, right=False)  # 按风速bin分组

    df_scada_WT = df_scada_WT.dropna()  # 删除无法归到bin里的空值
    df_scada_WT = df_scada_WT.reset_index(drop=True)
    df_scada_WT_grouped = df_scada_WT.groupby(df_scada_WT['windSpeed_lable'])  # 按风速bin分组

    powerCurve01 = powerCurve()
    powerCurve01.turbineName = df_scada_WT.loc[0, 'turbineName']

    list_eachbin_df = []
    for name, group in df_scada_WT_grouped:
        if (len(group) > 0):
            df_temp = df_scada_WT_grouped.get_group(name)  # get_group 才能返回dataFrame, group 是tuple
            list_eachbin_df.append(df_temp)

    nacellVib_Y_ave = pd.DataFrame()
    for i_bin in range(len(list_eachbin_df)):  # 对每一个风速bin，进行数据清洗 , power 四分位法

        res_var = np.percentile(list_eachbin_df[i_bin]['nacelleVib_Y'], (25, 50, 75), interpolation='midpoint')
        Iqr_var = res_var[2] - res_var[0]
        var_l = res_var[0] - 0.001 * Iqr_var
        var_u = res_var[2] + 0.001 * Iqr_var
        if (len(list_eachbin_df[i_bin]) > 5):
            list_eachbin_df[i_bin] = list_eachbin_df[i_bin][
                (list_eachbin_df[i_bin][vib_y] >= var_l) & (list_eachbin_df[i_bin][vib_x] <= var_u)]

        new = pd.DataFrame({'windSpeed_ave': np.mean(list_eachbin_df[i_bin][ws]), 'Y_ave':np.mean(list_eachbin_df[i_bin][vib_y])}, index = ['0'])

        nacellVib_Y_ave = nacellVib_Y_ave.append(new,ignore_index=True)

    df_scada_WT_1['nacellVib_Y_shift']=np.mean(nacellVib_Y_ave['Y_ave'])

def shiftVib (list_df_WTs):
    for i in range(len(list_df_WTs)):
        list_df_WTs[i] = list_df_WTs[i][list_df_WTs[i]['nacelleVib_X'] > -0.1]
        list_df_WTs[i] = list_df_WTs[i][list_df_WTs[i]['nacelleVib_X'] < 0.1]


        list_df_WTs[i] = list_df_WTs[i].reset_index(drop=True)
        #####################################

        vib_shift_X(list_df_WTs[i])
        vib_shift_Y(list_df_WTs[i])
        list_df_WTs[i]['nacelleVib_X_new'] = list_df_WTs[i]['nacelleVib_X'] - list_df_WTs[i]['nacellVib_X_shift']
        list_df_WTs[i]['nacelleVib_Y_new'] = list_df_WTs[i]['nacelleVib_Y'] - list_df_WTs[i]['nacellVib_Y_shift']
        list_df_WTs[i]['vib_magn_new'] = list(map(lambda x, y: virb_magn(x, y),
                                                  list_df_WTs[i]['nacelleVib_X_new'],
                                                  list_df_WTs[i]['nacelleVib_Y_new']))

    print(i, 'Vib_shifted')
def cleanPower(list_df_WTs,R,low_Cp,low_power):
    for i in range(len(list_df_WTs)):
        list_df_WTs[i]['powerLow']=1/2*1.225*math.pi*R**2*list_df_WTs[i][ws]**3*low_Cp/1000.0
        list_df_WTs[i]['powerLow'][list_df_WTs[i]['powerLow']>low_power]=low_power
        list_df_WTs[i]=list_df_WTs[i][list_df_WTs[i][gp]>list_df_WTs[i]['powerLow']]

def cal_Rho(list_df_WTs):
    for i in range(len(list_df_WTs)):

        list_df_WTs[i]['Ps'] = np.exp( 6.42 + 0.072 * list_df_WTs[i]['temperature'] - 0.000271 * list_df_WTs[i]['temperature']**2 + 0.000000723 * list_df_WTs[i]['temperature']**3)

        if  ('pressure' not in list_df_WTs[i].columns):

            list_df_WTs[i]['pressure']=101525.0
        if  ('humidity' not in list_df_WTs[i].columns):
            list_df_WTs[i]['humidity']=1.0

        list_df_WTs[i]['Rho_standard']=0.003483*list_df_WTs[i]['pressure']/(list_df_WTs[i]['temperature']+273)*(1-0.3779*list_df_WTs[i]['humidity'] *list_df_WTs[i]['Ps'] /list_df_WTs[i]['pressure'] )

def cal_Cp(list_df_WTs,R,yawCorrect, rhoCorrect):
    for i in range(len(list_df_WTs)):
        if (yawCorrect==0):
            list_df_WTs[i]['yawError']=0
        if (rhoCorrect == 0):
            list_df_WTs[i]['Rho_standard']=1.225

        # list_df_WTs[i][ws] = list_df_WTs[i][ws] * 0.8904 + 0.7937 #机舱传递函数
        list_df_WTs[i]['windSpeed_yawCorrected'] = list_df_WTs[i][ws] * np.cos(list_df_WTs[i][yaw] * math.pi / 180)
        list_df_WTs[i]['windSpeed_Standard'] = list_df_WTs[i]['windSpeed_yawCorrected'] * ((list_df_WTs[i]['Rho_standard'] / 1.225) ** (1 / 3))
        list_df_WTs[i]['Cp']=list_df_WTs[i][gp]*1000/(1/2*1.225*math.pi*R**2*list_df_WTs[i]['windSpeed_Standard']**3)
        list_df_WTs[i]['Cp_noRhoCorrect'] = list_df_WTs[i][gp] * 1000 / (1 / 2 * 1.225 * math.pi * R ** 2 * list_df_WTs[i]['windSpeed_yawCorrected'] ** 3)

def get_powerCurve (df_scada_WT, windType, delta_bin, V_in, V_out,clean_low, clean_up,R ): ########功率曲线处理函数，
    # 定义功率曲线处理配置参数
    bin_Num =int((V_out-V_in)/delta_bin)
    windSpeedBins = np.linspace(V_in, V_out, bin_Num+1)
    windSpeedLables = np.linspace(V_in+delta_bin/2, V_out-delta_bin/2, bin_Num)

    # df_scada_WT = df_scada_WT.loc[(df_scada_WT[rs] > rs_lower)
    #           & (df_scada_WT[ws] > wsin)
    #           & (df_scada_WT[gp] > gp_lower)
    #           & (((df_scada_WT[gp] < maxgp * gp_percent_limit) & (df_scada_WT[bp] < ba_limit))
    #              | ((df_scada_WT[gp] >= maxgp * gp_percent_limit) & (df_scada_WT[bp] <= ba_upper))) ]

    df_scada_WT['windSpeed_lable'] = pd.cut(df_scada_WT[windType], bins= windSpeedBins, labels=windSpeedLables, right =False) # 按风速bin分组

    df_scada_WT =df_scada_WT.dropna()#删除无法归到bin里的空值
    df_scada_WT= df_scada_WT.reset_index(drop=True)
    df_scada_WT_grouped =df_scada_WT.groupby(df_scada_WT['windSpeed_lable']) # 按风速bin分组

    powerCurve01 = powerCurve()
    powerCurve01.turbineName = df_scada_WT.loc[0,'turbineName']

    list_eachbin_df=[]
    for name,group in df_scada_WT_grouped:
        if (len(group)>0):
            df_temp= df_scada_WT_grouped.get_group(name) #get_group 才能返回dataFrame, group 是tuple
            list_eachbin_df.append(df_temp)

    for i in range(len(list_eachbin_df)): # 对每一个风速bin，进行数据清洗 , power 四分位法
        #功率清洗
        try:
            res_power = np.percentile(list_eachbin_df[i]['power'], (clean_low, (clean_low+clean_up)/2, clean_up), interpolation='midpoint')
            Iqr_power = res_power[2]-res_power[0]
            power_l= res_power[0]-0.001*Iqr_power
            power_u =res_power[2] +0.001*Iqr_power
            if(len(list_eachbin_df[i])>5):
                list_eachbin_df[i]=list_eachbin_df[i][(list_eachbin_df[i]['power']>=power_l) & (list_eachbin_df[i]['power']<=power_u)]

            if (len(list_eachbin_df[i]) > 0):
                list_eachbin_df[i] = list_eachbin_df[i].reset_index(drop=True)

                # 新增给powerCurve01.eachbin_stats_data
                new = pd.DataFrame({"V_average_processed": np.mean(list_eachbin_df[i][windType]),
                                    "P_average_processed": np.mean(list_eachbin_df[i]['power']),
                                    "V_l": list_eachbin_df[i].loc[0, 'windSpeed_lable'] - 0.5*delta_bin,
                                    "V_u": list_eachbin_df[i].loc[0, 'windSpeed_lable'] + 0.5*delta_bin,
                                    "V_label": list_eachbin_df[i].loc[0, 'windSpeed_lable'],
                                    "Cp_ave": np.mean(list_eachbin_df[i]['power'])*1000/(1/2*1.225*math.pi*R**2*np.mean(list_eachbin_df[i][windType])**3),
                                    # 'windDirection_ave':np.mean(list_eachbin_df[i]['windDirection'])
                                    }, index=["0"])

                powerCurve01.eachbin_stats_data = powerCurve01.eachbin_stats_data.append(new, ignore_index=True)


        except:
            for j in range(len(list_eachbin_df[i])):
                print ('something wrong when getting power curve of', i, j)

    powerCurve01.list_eachbin_df= list_eachbin_df
    return powerCurve01

def get_vibCurve (df_scada_WT, windType, delta_bin, V_in, V_out,clean_low, clean_up): ########功率曲线处理函数，
    # 定义功率曲线处理配置参数
    bin_Num =int((V_out-V_in)/delta_bin)
    windSpeedBins = np.linspace(V_in, V_out, bin_Num+1)
    windSpeedLables = np.linspace(V_in+delta_bin/2, V_out-delta_bin/2, bin_Num)

    df_scada_WT['windSpeed_lable'] = pd.cut(df_scada_WT[windType], bins= windSpeedBins, labels=windSpeedLables, right =False) # 按风速bin分组

    df_scada_WT =df_scada_WT.dropna()#删除无法归到bin里的空值
    df_scada_WT= df_scada_WT.reset_index(drop=True)
    df_scada_WT_grouped =df_scada_WT.groupby(df_scada_WT['windSpeed_lable']) # 按风速bin分组

    vibCurve01 = vibCurve()
    vibCurve01.turbineName = df_scada_WT.loc[0,'turbineName']

    list_eachbin_df=[]
    for name,group in df_scada_WT_grouped:
        if (len(group)>0):
            df_temp= df_scada_WT_grouped.get_group(name) #get_group 才能返回dataFrame, group 是tuple
            list_eachbin_df.append(df_temp)

    for i in range(len(list_eachbin_df)): # 对每一个风速bin，进行数据清洗 , power 四分位法

        try:
            # 振动清洗
            res_virb = np.percentile(list_eachbin_df[i]['vib_magn_new'], (clean_low, (clean_low+clean_up)/2, clean_up), interpolation='midpoint')  # lambda 四分位法
            Iqr_virb = res_virb[2] - res_virb[0]
            virb_l = res_virb[0] - 0.001 * Iqr_virb
            virb_u = res_virb[2] + 0.001 * Iqr_virb
            if (len(list_eachbin_df[i]) > 5):
                list_eachbin_df[i] = list_eachbin_df[i][
                    (list_eachbin_df[i]['vib_magn_new'] >= virb_l) & (list_eachbin_df[i]['vib_magn_new'] <= virb_u)]


            if (len(list_eachbin_df[i]) > 0):
                list_eachbin_df[i] = list_eachbin_df[i].reset_index(drop=True)
                # 新增给powerCurve01.eachbin_stats_data
                new = pd.DataFrame({"V_average_processed": np.mean(list_eachbin_df[i][windType]),
                                    "V_l": list_eachbin_df[i].loc[0, 'windSpeed_lable'] - 0.5*delta_bin,
                                    "V_u": list_eachbin_df[i].loc[0, 'windSpeed_lable'] + 0.5*delta_bin,
                                    "V_label": list_eachbin_df[i].loc[0, 'windSpeed_lable'],
                                    "vib_magn_average_processed": np.mean(list_eachbin_df[i]['vib_magn_new'])

                                    }, index=["0"])
                vibCurve01.eachbin_stats_data = vibCurve01.eachbin_stats_data.append(new, ignore_index=True)
        except:
            for j in range(len(list_eachbin_df[i])):
                print ('something wrong when getting vibration curve of', i, j)

    vibCurve01.list_eachbin_df= list_eachbin_df
    return vibCurve01


def get_AEP (eachbin_stats_data: pd.DataFrame, V_average, weibull_k, Vrate, Prate):
    AEP =0
    for i in range(len(eachbin_stats_data)):
        weibull_a = V_average/math.exp(math.lgamma(1/weibull_k+1))
        CPD_l = weibull_min.cdf(eachbin_stats_data.loc[i, 'V_l']/weibull_a, weibull_k)
        CPD_u = weibull_min.cdf(eachbin_stats_data.loc[i, 'V_u'] / weibull_a, weibull_k)
        prob = CPD_u- CPD_l
        if eachbin_stats_data.loc[i, 'V_u'] <= Vrate:
            delt_power = eachbin_stats_data.loc[i, 'P_average_processed']*prob *8766
        else:
            delt_power = Prate * prob * 8766
        AEP = AEP+ delt_power
    return AEP

def print_Vars(savePath, sheet, rowStart, turbineName, AEP, Hours, CF, vib_overall):

    wb = openpyxl.load_workbook(savePath)
    ws = wb[sheet]
    ws.cell(row=rowStart, column=2).value = turbineName
    ws.cell(row=rowStart, column=3).value = AEP
    ws.cell(row=rowStart, column=4).value = Hours
    ws.cell(row=rowStart, column=5).value = CF
    ws.cell(row=rowStart, column=6).value = vib_overall

    wb.save(savePath)

def cal_AEP (list_df_WTs, output_Dir, bin, Vin, Vout, Vrate, Prate, R, V_average, weibull_k, windType, clean_low, clean_up,bymonth,in_date, subNumber_H, subNumber_W, cut_lead, cut_tail,  ifShow):
    print ('Wait for Calculating Power Curve...')
    # savePath = output_Dir + str(list_df_WTs[i].loc[0]['turbineName']) + r'\\'
    savePath = output_Dir
    if not os.path.exists(savePath):
        os.makedirs(savePath)
    filePath = savePath  + 'PowerCurve.xlsx'
    if os.path.isfile(filePath):
        os.remove(filePath)
    if not os.path.isfile(filePath):
        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        wb.create_sheet("PU")
        wb.create_sheet("AEP")
        wb.save(filename=filePath)
    plt.figure(figsize=(16, 16))
    for i in range (cut_lead, (len(list_df_WTs)-cut_tail)):
        plt.subplot(subNumber_H, subNumber_W, i + 1)

        powerCurve_one = get_powerCurve(list_df_WTs[i], R=R,  windType=windType, delta_bin=bin, V_in=Vin, V_out=Vout, clean_low =clean_low, clean_up=clean_up)
        AEP = get_AEP(powerCurve_one.eachbin_stats_data, V_average=V_average, weibull_k=weibull_k, Vrate=Vrate, Prate=Prate)
        Hours = AEP / Prate
        CF = Hours / 8766
        print('PowerCurve got', powerCurve_one.turbineName, 'AEP(kWh)=', AEP, 'Hours(h) =', Hours, 'CF(-)=', CF)
        if (bymonth==0):
            powerCurve_one.print_to_excel(filePath,'PU', 1+1*i,delta_bin=bin,V_in=Vin, V_out=Vout)

        if (bymonth==1):

            dt = datetime.datetime.strptime(in_date, "%Y-%m")
            out_date = (dt + relativedelta(months=+i)).strftime("%Y-%m")
            powerCurve_one.print_to_excel_byMonth(filePath,'PU', 1+1*i, startMonth=out_date,delta_bin=bin,V_in=Vin, V_out=Vout)

        print_Vars(savePath=filePath, sheet='AEP',rowStart= i + 2, turbineName=powerCurve_one.turbineName, AEP=AEP, Hours=Hours, CF=CF,vib_overall=0)

        plt.scatter(list_df_WTs[i][windType], list_df_WTs[i]['power'], color='b', label='origin scada', s=0.1)
        for j in range (len(powerCurve_one.list_eachbin_df)):
            if j==1 :
                plt.scatter(powerCurve_one.list_eachbin_df[j][windType], powerCurve_one.list_eachbin_df[j]['power'], label='processed scada', color='tomato', s=1)
            else :
                plt.scatter(powerCurve_one.list_eachbin_df[j][windType], powerCurve_one.list_eachbin_df[j]['power'], color='tomato', s=1)

            plt.plot (powerCurve_one.eachbin_stats_data['V_average_processed'], powerCurve_one.eachbin_stats_data['P_average_processed'], color='black', label='processed_mean')
            plt.ylim((0, Prate*1.1))
        plt.xlabel('WindSpeed m/s')
        plt.ylabel('Power kW')
        plt.xlim((0, 20))

        plt.title(str(powerCurve_one.turbineName), fontsize='medium', y=0.1)
        plt.tick_params(labelsize=10)
        plt.subplots_adjust(left=0.2, top=0.96, right=0.96, bottom=0.1, wspace=0.3, hspace=0.3)


    plt.savefig(savePath  + windType + 'PowerCurves.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
    if (ifShow == True):
        plt.show()
    plt.clf()


    if (bymonth == 0):
        plt.figure(figsize=(16, 16))


        # NUM_COLORS = len(list_df_WTs) - cut_tail
        # cm = plt.get_cmap('gist_rainbow')
        # fig = plt.figure()
        # ax = fig.add_subplot(111)
        # ax.set_prop_cycle('color', [cm(1. * i / NUM_COLORS) for i in range(NUM_COLORS)])

        for i in range(cut_lead, (len(list_df_WTs) - cut_tail)):
            powerCurve_one = get_powerCurve(list_df_WTs[i], R=R, windType=windType, delta_bin=bin, V_in=Vin, V_out=Vout,  clean_low=clean_low, clean_up=clean_up)
            plt.plot(powerCurve_one.eachbin_stats_data['V_average_processed'],powerCurve_one.eachbin_stats_data['P_average_processed'], label=powerCurve_one.turbineName)

            # ax.plot(powerCurve_one.eachbin_stats_data['V_average_processed'], powerCurve_one.eachbin_stats_data['P_average_processed'],  label=powerCurve_one.turbineName )
        plt.legend(fontsize= 15)
        plt.ylim((0, Prate*1.1))
        plt.xlabel('WindSpeed m/s')
        plt.ylabel('Power kW')
        plt.xlim((0, 20))
        # plt.show()
        plt.savefig(savePath + 'All_turbine_PowerCurves.jpg', format='jpg',dpi=plt.gcf().dpi, bbox_inches='tight')

        plt.clf()



def cal_Vibration (list_df_WTs, output_Dir, bin, Vin, Vout,  windType, clean_low, clean_up, subNumber_H, subNumber_W, cut_lead, cut_tail,  ifShow):
    print ('Wait for Calculating Vibration...')
    # savePath = output_Dir + str(list_df_WTs[i].loc[0]['turbineName']) + r'\\'
    savePath = output_Dir
    if not os.path.exists(savePath):
        os.makedirs(savePath)
    filePath = savePath  + 'Vibration.xlsx'
    if os.path.isfile(filePath):
        os.remove(filePath)
    if not os.path.isfile(filePath):
        # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
        wb = Workbook()
        wb.create_sheet("Vib-U")
        wb.create_sheet("Sum")
        wb.save(filename=filePath)

    for i in range (cut_lead, (len(list_df_WTs)-cut_tail)):
        plt.subplot(subNumber_H, subNumber_W, i + 1)
        vibCurve_one = get_vibCurve(list_df_WTs[i], windType=windType, delta_bin=bin, V_in=Vin, V_out=Vout, clean_low=clean_low, clean_up=clean_up)

        print('Vibration got', vibCurve_one.turbineName, 'vir =', np.mean(vibCurve_one.eachbin_stats_data['vib_magn_average_processed']))

        vibCurve_one.print_to_excel(filePath, 'Vib-U', 1 + 1 * i, delta_bin=bin, V_in=Vin, V_out=Vout)

        print_Vars(savePath=filePath, sheet='Sum', rowStart=i + 2, turbineName=vibCurve_one.turbineName, AEP=0, Hours=0, CF=0, vib_overall=np.mean(vibCurve_one.eachbin_stats_data['vib_magn_average_processed']))

        plt.scatter(list_df_WTs[i][windType], list_df_WTs[i]['vib_magn_new'], color='b', label='origin scada', s=0.1)
        for j in range(len(vibCurve_one.list_eachbin_df)):
            if j == 1:
                plt.scatter(vibCurve_one.list_eachbin_df[j][windType], vibCurve_one.list_eachbin_df[j]['vib_magn_new'],
                            label='processed scada', color='tomato', s=1)
            else:
                plt.scatter(vibCurve_one.list_eachbin_df[j][windType], vibCurve_one.list_eachbin_df[j]['vib_magn_new'],
                            color='tomato', s=1)

            plt.plot(vibCurve_one.eachbin_stats_data['V_average_processed'],
                     vibCurve_one.eachbin_stats_data['vib_magn_average_processed'], color='black', label='processed_mean')
            plt.ylim((0, 0.05))

        plt.xlim((0, 20))

        # plt.title('Turbine_' + str(vibCurve_one.turbineName), fontsize='medium', y=0.5)
        plt.xlabel('风速 m/s')
        plt.ylabel('加速度 g')
        plt.tick_params(labelsize=10)
        plt.subplots_adjust(left=0.2, top=0.96, right=0.96, bottom=0.1, wspace=0.3, hspace=0.3)

    plt.savefig(savePath + windType + 'VibrationCurves.jpg', format='jpg', dpi=plt.gcf().dpi, bbox_inches='tight')
    if (ifShow == True):
        plt.show()
    plt.clf()

def dividedIntoMonths (list_df_WTs, output_Dir, bin, Vin, Vout, Vrate, Prate, R, V_average, weibull_k, windType, bymonth,in_date, clean_low, clean_up, subNumber_H, subNumber_W, cut_lead, cut_tail, vibExist, ifShow):

    for i in range(len(list_df_WTs)):
        list_df_WTs[i] = list_df_WTs[i].drop_duplicates(['time'])
        list_df_WTs[i].set_index(["time"], inplace=True)

        df_scada_WT_groupedbyTime = list_df_WTs[i].groupby(pd.Grouper(freq='M'))

        list_eachbin_df = []
        for name, group in df_scada_WT_groupedbyTime:
            if (len(group) > 0):
                df_temp = df_scada_WT_groupedbyTime.get_group(name)  # get_group 才能返回dataFrame, group 是tuple
                list_eachbin_df.append(df_temp)

        for j in range(len(list_eachbin_df)):
            saveDataPath = output_Dir + 'dividedByMonth\\'+ list_eachbin_df[j]['turbineName'][0]+'\\Data_Raw'
            filePath = output_Dir + 'dividedByMonth\\' + list_eachbin_df[j]['turbineName'][0]
            if not os.path.exists(saveDataPath):
                os.makedirs(saveDataPath)

            path = saveDataPath + '\\'+list_eachbin_df[j]['turbineName'][0] + '_' + str(j).zfill(2) + '.csv'

            list_eachbin_df[j]['turbineName'] = list_eachbin_df[j]['turbineName'][0] + '_' + str(j).zfill(2)
            list_eachbin_df[j].to_csv(path, encoding='gb2312')

        cal_AEP(list_eachbin_df, output_Dir=filePath+'\\', bin=bin, Vin=Vin, Vout=Vout, R=R, Vrate=Vrate, Prate=Prate,
                V_average=V_average, weibull_k=weibull_k, windType=windType,bymonth=bymonth,in_date=in_date,
                clean_low=clean_low, clean_up=clean_up,
                cut_lead=cut_lead, cut_tail=cut_tail, subNumber_H=subNumber_H, subNumber_W=subNumber_W, ifShow=ifShow)

        if (vibExist==True):
            cal_Vibration(list_eachbin_df, output_Dir=filePath+'\\', bin=bin, Vin=Vin, Vout=Vout, windType=ws,
                          clean_low=80, clean_up=90,
                          cut_lead=cut_lead, cut_tail=cut_tail, subNumber_H=subNumber_H, subNumber_W=subNumber_W, ifShow=ifShow)